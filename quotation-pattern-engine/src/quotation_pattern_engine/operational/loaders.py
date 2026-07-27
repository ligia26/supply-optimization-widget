from __future__ import annotations

import csv
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from .config import OperationalConfig
from .models import DemandProfile, PatternEvent, PatternSummary, QuotationPoint, TankState


def _as_float(value: object, default: float = 0.0) -> float:
    if value is None or value == "":
        return default
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace("€", "").replace(",", "").strip()
    try:
        return float(text)
    except ValueError:
        return default


def load_tank_states(path: str | Path, config: OperationalConfig) -> list[TankState]:
    workbook = load_workbook(path, data_only=True, read_only=False)
    states: list[TankState] = []

    for sheet in workbook.worksheets:
        distributor_id = str(sheet.title).strip()
        tank_totals: dict[str, dict[str, float]] = defaultdict(
            lambda: {"capacity": 0.0, "inventory": 0.0}
        )

        for row in sheet.iter_rows(min_row=2, max_row=12, max_col=6, values_only=True):
            if not row:
                continue
            product = str(row[0]).strip() if row[0] is not None else ""
            if not product or product in config.ignored_products:
                continue
            capacity = _as_float(row[2])
            inventory = _as_float(row[3])
            if capacity <= 0:
                continue
            tank_totals[product]["capacity"] += capacity
            tank_totals[product]["inventory"] += inventory

        commercial: dict[str, tuple[float, float | None]] = {}
        for row in sheet.iter_rows(min_row=13, max_col=6, values_only=True):
            if not row:
                continue
            label = str(row[0]).strip() if row[0] is not None else ""
            if label not in {"Benzina", "Diesel"}:
                continue
            purchase_price = _as_float(row[1])
            selling_price = _as_float(row[2], default=0.0) or None
            operational_product = "Verde" if label == "Benzina" else "Diesel"
            commercial[operational_product] = (purchase_price, selling_price)

        for product, totals in tank_totals.items():
            purchase_price, selling_price = commercial.get(product, (0.0, None))
            if purchase_price <= 0:
                raise ValueError(
                    f"Missing opening purchase/implicit cost for {distributor_id} {product}."
                )
            states.append(
                TankState(
                    distributor_id=distributor_id,
                    product=product,
                    capacity_litres=totals["capacity"],
                    opening_inventory_litres=totals["inventory"],
                    opening_implicit_cost_per_litre=purchase_price,
                    selling_price_per_litre=selling_price,
                )
            )

    return sorted(states, key=lambda item: (item.distributor_id, item.product))


def load_daily_sales(path: str | Path, config: OperationalConfig) -> list[dict[str, object]]:
    workbook = load_workbook(path, data_only=True, read_only=False)
    records: list[dict[str, object]] = []

    for sheet in workbook.worksheets:
        if not sheet.title.endswith("2025"):
            continue
        month_start = datetime.strptime(sheet.title, "%B %Y").date().replace(day=1)
        distributor_headers: list[tuple[int, str, str]] = []
        current_distributor = ""

        # Row 2 carries distributor names; row 3 carries product names.
        for col in range(3, sheet.max_column + 1):
            distributor_name = sheet.cell(2, col).value
            product = sheet.cell(3, col).value
            if distributor_name is not None:
                current_distributor = str(distributor_name).strip()
            if product in {"verde", "diesel"}:
                distributor_id = config.distributor_mapping.get(current_distributor)
                if distributor_id:
                    distributor_headers.append((col, distributor_id, str(product).title()))

        for row in range(4, sheet.max_row + 1):
            day_value = sheet.cell(row, 1).value
            if not isinstance(day_value, (int, float)):
                continue
            day = int(day_value)
            try:
                record_date = date(month_start.year, month_start.month, day)
            except ValueError:
                continue

            for col, distributor_id, product in distributor_headers:
                value = sheet.cell(row, col).value
                if value is None:
                    continue
                litres = _as_float(value)
                # Zeros in the source are retained: they may represent closure/no sales.
                records.append(
                    {
                        "date": record_date,
                        "weekday": record_date.weekday(),
                        "distributor_id": distributor_id,
                        "product": product,
                        "litres": litres,
                    }
                )

    return records


def load_growth_factors(path: str | Path, config: OperationalConfig) -> dict[str, float]:
    workbook = load_workbook(path, data_only=True, read_only=False)
    sheet = workbook.active

    names_2025 = [str(sheet.cell(4, col).value).strip().lower() for col in range(3, 8)]
    names_2026 = [str(sheet.cell(4, col).value).strip().lower() for col in range(8, 13)]

    totals_2025: dict[str, float] = defaultdict(float)
    totals_2026: dict[str, float] = defaultdict(float)

    # Jan-Jun only, because 2026 contains Jan-Jun in the supplied workbook.
    for row in range(5, 11):
        for offset, name in enumerate(names_2025, start=3):
            distributor_id = config.monthly_name_mapping.get(name)
            if distributor_id:
                totals_2025[distributor_id] += _as_float(sheet.cell(row, offset).value)
        for offset, name in enumerate(names_2026, start=8):
            distributor_id = config.monthly_name_mapping.get(name)
            if distributor_id:
                totals_2026[distributor_id] += _as_float(sheet.cell(row, offset).value)

    factors: dict[str, float] = {}
    for distributor_id in set(totals_2025) | set(totals_2026):
        base = totals_2025.get(distributor_id, 0.0)
        current = totals_2026.get(distributor_id, 0.0)
        factors[distributor_id] = current / base if base > 0 and current > 0 else 1.0
    return factors


def build_demand_profiles(
    daily_sales: Iterable[dict[str, object]],
    growth_factors: dict[str, float],
) -> list[DemandProfile]:
    grouped: dict[tuple[str, str, int], list[float]] = defaultdict(list)
    for row in daily_sales:
        key = (
            str(row["distributor_id"]),
            str(row["product"]),
            int(row["weekday"]),
        )
        grouped[key].append(float(row["litres"]))

    profiles: list[DemandProfile] = []
    for (distributor_id, product, weekday), values in grouped.items():
        average = sum(values) / len(values) if values else 0.0
        profiles.append(
            DemandProfile(
                distributor_id=distributor_id,
                product=product,
                weekday=weekday,
                average_daily_litres_2025=average,
                growth_factor_2026=growth_factors.get(distributor_id, 1.0),
            )
        )
    return sorted(profiles, key=lambda p: (p.distributor_id, p.product, p.weekday))


def load_quotation_points(
    daily_analysis_csv: str | Path,
    config: OperationalConfig,
) -> list[QuotationPoint]:
    reverse_mapping = {value: key for key, value in config.product_mapping.items()}
    rows: list[QuotationPoint] = []
    with Path(daily_analysis_csv).open(newline="", encoding="utf-8") as handle:
        for raw in csv.DictReader(handle):
            quotation_product = raw["product"].strip()
            product = reverse_mapping.get(quotation_product)
            if not product:
                continue
            change = raw.get("change")
            rows.append(
                QuotationPoint(
                    date=date.fromisoformat(raw["date"]),
                    product=product,
                    quotation_product=quotation_product,
                    price_per_litre=float(raw["price"]) / config.quotation_price_divisor,
                    regime=raw.get("regime", "Stable") or "Stable",
                    event_type=raw.get("event_type", "Normal quotation") or "Normal quotation",
                    change_per_litre=(
                        float(change) / config.quotation_price_divisor
                        if change not in (None, "")
                        else None
                    ),
                )
            )
    return sorted(rows, key=lambda item: (item.date, item.product))



def _first(raw: dict[str, str], *names: str, default: str = "") -> str:
    normalized = {str(k).strip().lower(): ("" if v is None else str(v).strip()) for k, v in raw.items()}
    for name in names:
        value = normalized.get(name.lower(), "")
        if value != "":
            return value
    return default


def _parse_date_flexible(value: str) -> date:
    value = value.strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Unsupported date value: {value!r}")


def _optional_float(value: str, divisor: float = 1.0) -> float | None:
    if value in (None, ""):
        return None
    cleaned = str(value).replace("€", "").replace(",", "").strip()
    try:
        return float(cleaned) / divisor
    except ValueError:
        return None


def load_pattern_events(
    path: str | Path,
    config: OperationalConfig,
) -> list[PatternEvent]:
    """Load pattern_events.csv while tolerating minor column-name differences."""
    reverse_mapping = {value: key for key, value in config.product_mapping.items()}
    events: list[PatternEvent] = []
    with Path(path).open(newline="", encoding="utf-8") as handle:
        for raw in csv.DictReader(handle):
            quotation_product = _first(raw, "product", "quotation_product")
            product = reverse_mapping.get(quotation_product, quotation_product)
            if product not in config.product_mapping:
                continue
            pattern_id = _first(raw, "pattern_id", "pattern", "id")
            pattern_name = _first(raw, "pattern_name", "name", "event_name", "description", default=pattern_id)
            start_text = _first(raw, "start_date", "date", "from_date")
            end_text = _first(raw, "end_date", "date", "to_date", default=start_text)
            if not start_text:
                continue
            magnitude = _optional_float(
                _first(raw, "magnitude", "total_change", "change", "value"),
                config.quotation_price_divisor,
            )
            threshold = _optional_float(
                _first(raw, "threshold", "spike_threshold"),
                config.quotation_price_divisor,
            )
            duration_text = _first(raw, "duration_days", "duration", "length", "streak_length")
            duration = int(float(duration_text)) if duration_text else None
            confidence_text = _first(
                raw,
                "confidence",
                "confidence_level",
                "pattern_confidence",
                "detection_confidence",
                "score",
                "probability",
            )
            confidence_raw = _optional_float(confidence_text)
            if confidence_raw is None and confidence_text:
                confidence_raw = config.confidence_labels.get(confidence_text.strip().lower())
            if confidence_raw is not None and confidence_raw > 1.0:
                confidence_raw /= 100.0
            if confidence_raw is not None:
                confidence_raw = min(1.0, max(0.0, confidence_raw))

            events.append(PatternEvent(
                product=product,
                pattern_id=pattern_id,
                pattern_name=pattern_name,
                start_date=_parse_date_flexible(start_text),
                end_date=_parse_date_flexible(end_text),
                direction=_first(raw, "direction", "trend", "regime"),
                magnitude_per_litre=magnitude,
                threshold_per_litre=threshold,
                duration_days=duration,
                event_type=_first(raw, "event_type", "type", "category"),
                confidence=confidence_raw,
                raw={str(k): "" if v is None else str(v) for k, v in raw.items()},
            ))
    return sorted(events, key=lambda e: (e.product, e.start_date, e.end_date, e.pattern_id))


def load_pattern_summaries(path: str | Path, config: OperationalConfig) -> list[PatternSummary]:
    """Load product-level summary metrics; nonnumeric fields are retained."""
    summaries: list[PatternSummary] = []
    reverse_mapping = {value: key for key, value in config.product_mapping.items()}
    with Path(path).open(newline="", encoding="utf-8") as handle:
        for raw in csv.DictReader(handle):
            quotation_product = _first(raw, "product", "quotation_product")
            product = reverse_mapping.get(quotation_product, quotation_product)
            metrics: dict[str, float | str] = {}
            for key, value in raw.items():
                if key is None or key.lower() in {"product", "quotation_product"}:
                    continue
                text = "" if value is None else str(value).strip()
                numeric = _optional_float(text)
                metrics[key] = numeric if numeric is not None else text
            summaries.append(PatternSummary(product=product, metrics=metrics))
    return summaries
